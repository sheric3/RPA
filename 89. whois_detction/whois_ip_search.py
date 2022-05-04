# -*- coding: utf-8 -*-
import json
import requests
from openpyxl import Workbook
from openpyxl import load_workbook
# from openpyxl.compat import range

wb = Workbook()

# 엑셀의 액티브 워크시트 선택
ws = wb.active
ws.title = "whois ip info"
column_num = 2


# 엑셀에 컬럼 저장
def saveColName(cName):
    column_char = 'a'
    for name in cName:
        ws[column_char + '1'] = name
        column_char = chr(ord(column_char) + 1)


# 엑셀에 조회결과 저장
def saveContent(cName):
    column_char = 'a'
    global column_num
    for name in cName:
        ws[column_char + str(column_num)] = name
        column_char = chr(ord(column_char) + 1)


# 세션 생성
s = requests.session()

def requestWhois(searchIP):
    # 엑셀 2번째 줄부터 IP정보를 연달아 저장하기 위해서 전역 변수 선언
    global column_num

    # API 호출 하기(발급받은키는 자신의 키로 수정해야 한다)
    con = s.get('http://whois.kisa.or.kr/openapi/whois.jsp?query=' + searchIP + '&key=2022041510472622010986&answer=json')

    # 호출된 결과를 json 형태로 저장
    json_data = json.loads(con.text)
    # 결과를 담을 배열 초기화
    WhoIsData = []

    # 각 결과 값을 배열에 저장
    WhoIsData.append(json_data['whois']['query'])
    WhoIsData.append(json_data['whois']['countryCode'])

    # json 인자가 있는지 체크하는 식으로 에러처리 바꿈
    if 'korean' in json_data['whois']:
        if 'PI' in json_data['whois']['korean']:
            WhoIsData.append(json_data['whois']['korean']['PI']['netinfo']['addr'])
            WhoIsData.append(json_data['whois']['korean']['PI']['netinfo']['range'])
            WhoIsData.append(json_data['whois']['korean']['PI']['netinfo']['servName'])
        elif 'ISP' in json_data['whois']['korean']:
            WhoIsData.append(json_data['whois']['korean']['ISP']['netinfo']['addr'])
            WhoIsData.append(json_data['whois']['korean']['ISP']['netinfo']['range'])
            WhoIsData.append(json_data['whois']['korean']['ISP']['netinfo']['servName'])

    # 결과 화면에 뿌리기
    print(WhoIsData)

    # 엑셀에 결과 저장
    saveContent(WhoIsData)
    column_num = column_num + 1


# 엑셀 1째 줄에 데이터 제목 저장
excelTitle = ['IP', 'countryCode', 'addr', 'IP_range', 'servName']
saveColName(excelTitle)

# 파일 열어 리스트에 담기
f = open('ip.txt', 'r')
iplist = f.read().splitlines()

# 리스트에 담기 IP 정보 얻어 오기
for ip in iplist:
    requestWhois(ip)

# 엑셀 저장하기
wb.save("ipinfo.xlsx")