import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook # 파일 불러오기
import datetime
import os
import win32com.client as win32
import pycxcel as p


# # xls to xlsx
# fname = "WAPPLESLog_0526_0000.xls"
# excel = win32.gencache.EnsureDispatch('Excel.Application')
# wb = excel.Workbooks.Open(fname)

wb.SaveAs(fname+"x", FileFormat = 51) #FileFormat = 51 is for .xlsx extension
wb.Close() #FileFormat = 56 is for .xls extension
excel.Application.Quit()

# # wb = load_workbook("E:/리틀팍스 정보보안/^2022년/# WAF 탐지 로그 (로우데이터)/2022/원본/WAPPLESLog_0526_0000.xls") # .xls 파일에서 wb를 불러온다.
# wb = openpyxl.load_workbook('WAPPLESLog_0526_0000.xlsx')
# print(wb)
#
# ws = wb.active # 활성화된 Sheet
#
# # 작업할 경로 변수
# file_src = 'E:/리틀팍스 정보보안/^2022년/# WAF 탐지 로그 (로우데이터)/2022/원본'
#
# # 폴더로 이동하여 파일 리스트 추출
# file_list = os.listdir(file_src)
# cnt = 1 #이름 중독 방지 변수

# # datetime을 문자열로 변환 당일 날짜
# # %Y 4자리 년도 표시, %y 2자리 년도 표시, %m 월 2자리 표시, %d 일을 2자리 표시, %H 시 출력, %M 분 출력, %S 초 출력
# today_md = datetime.today().strftime("%m%d")
# print(today_md) # 결과 : 0527
